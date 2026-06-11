#!/usr/bin/env python3
"""
Vision Infra - Summary Report Generator
=======================================
Regenerates the two-sheet workbook ("Master" + "Summery report") from a new
raw Master logsheet, inheriting the EXACT formulas, structure, styles, column
widths, number formats, conditional formatting and auto-filter of the reference
file (final_fine.xlsx).

The "Summery report" sheet stays fully formula-driven off "Master":
  A  Sr No            =ROW()-1
  C  Make            =IFERROR(INDEX(Master!I:I,MATCH($B,Master!G:G,0)),"")
  D  Model           =IFERROR(INDEX(Master!J:J,MATCH($B,Master!G:G,0)),"")
  F  DEPARTMENT      =IFERROR(INDEX(Master!L:L,MATCH($B,Master!G:G,0)),"")
  J  MONTH/YEAR      =TEXT(K,"mmm-yyyy")
  K  START DATE      =IFERROR(MINIFS(Master!B:B,Master!G:G,$B,Master!E:E,$G),"")
  L  CLOSE DATE      =IFERROR(MAXIFS(Master!B:B,Master!G:G,$B,Master!E:E,$G),"")
  M  Total Days      {=IFERROR(ROWS(UNIQUE(FILTER(dates, asset*site))),0)}
  N  Free Days       {= ... *status="Free Asset"}
  O  IN TRANSIT      {= ... *status="Transit"}
  P  DEPLOYED        =M-N-O
  Q  BDWN            {= ... *status="Breakdown"}
  R  Idle            {= ... *status="Idle"}
  S  Working         {= ... *status="Working"}
  T  Billing         =R+S
  U  WORKED HOURS    =SUMIFS(Master!T:T,Master!G:G,$B,Master!E:E,$G)
Inputs (filled from the new data): B Asset Code, G Site Code, H Client, I Location.

Usage:
  python build_summary_report.py <new_master.xlsx|.xls|.csv> [reference.xlsx] [output.xlsx]
"""
import sys, os, shutil, subprocess, zipfile, datetime
import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
import xml.etree.ElementTree as ET

SUMMARY_SHEET = "Summery report"
MASTER_SHEET  = "Master"
EXCEL_EPOCH   = datetime.date(1899, 12, 30)

# ---------- helpers ----------
def read_any(path):
    if path.lower().endswith(".csv"):
        return pd.read_csv(path)
    if path.lower().endswith(".xls"):                       # legacy -> convert
        out = os.path.dirname(os.path.abspath(path)) or "."
        subprocess.run(["libreoffice","--headless","--convert-to","xlsx","--outdir",out,path],
                       check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        path = os.path.splitext(path)[0] + ".xlsx"
    return pd.read_excel(path, sheet_name=MASTER_SHEET) if MASTER_SHEET in \
           pd.ExcelFile(path).sheet_names else pd.read_excel(path)

def serial(d):
    if pd.isna(d): return None
    d = pd.Timestamp(d).date()
    return (d - EXCEL_EPOCH).days

# ---------- formula templates (verbatim from reference) ----------
def formulas(r, N):
    A = f"=ROW()-1"
    C = f'=IFERROR(INDEX(Master!I:I,MATCH($B{r},Master!G:G,0)),"")'
    D = f'=IFERROR(INDEX(Master!J:J,MATCH($B{r},Master!G:G,0)),"")'
    F = f'=IFERROR(INDEX(Master!L:L,MATCH($B{r},Master!G:G,0)),"")'
    J = f'=TEXT(K{r},"mmm-yyyy")'
    K = f'=IFERROR(_xlfn.MINIFS(Master!B:B,Master!G:G,$B{r},Master!E:E,$G{r}),"")'
    L = f'=IFERROR(_xlfn.MAXIFS(Master!B:B,Master!G:G,$B{r},Master!E:E,$G{r}),"")'
    P = f"=M{r}-N{r}-O{r}"
    T = f"=R{r}+S{r}"
    U = f"=SUMIFS(Master!T:T,Master!G:G,$B{r},Master!E:E,$G{r})"
    def arr(status=None):
        base = (f"(Master!$G$2:$G${N}=$B{r})*(Master!$E$2:$E${N}=$G{r})")
        if status:
            base += f'*(Master!$U$2:$U${N}="{status}")'
        return (f"=IFERROR(ROWS(_xlfn.UNIQUE(_xlfn._xlws.FILTER("
                f"Master!$B$2:$B${N},{base}))),0)")
    M = arr(); N_ = arr("Free Asset"); O = arr("Transit")
    Q = arr("Breakdown"); R = arr("Idle"); S = arr("Working")
    return dict(A=A,C=C,D=D,F=F,J=J,K=K,L=L,M=M,N=N_,O=O,P=P,Q=Q,R=R,S=S,T=T,U=U)

# ---------- python-side values (for cached <v>, matches Excel exactly) ----------
def compute_values(master, pairs, first_by_asset):
    m = master
    vals = {}
    for i,(asset,site) in enumerate(pairs, start=1):
        sub = m[(m["equipmentnumber"]==asset) & (m["sitecode"]==site)]
        dates = sub["date"].dropna()
        def dd(st=None):
            s = sub if st is None else sub[sub["status"]==st]
            return int(s["date"].dropna().nunique())
        M=dd(); N=dd("Free Asset"); O=dd("Transit"); Q=dd("Breakdown")
        R=dd("Idle"); S=dd("Working")
        kmin = dates.min() if len(dates) else None
        kmax = dates.max() if len(dates) else None
        mk,md,dept = first_by_asset.get(asset,("","",""))
        jtext = pd.Timestamp(kmin).strftime("%b-%Y") if kmin is not None and not pd.isna(kmin) else ""
        vals[i] = dict(
            A=i, C=str(mk or ""), D=str(md or ""), F=str(dept or ""),
            J=jtext, K=serial(kmin), L=serial(kmax),
            M=M, N=N, O=O, P=M-N-O, Q=Q, R=R, S=S, T=R+S,
            U=round(float(sub["running_hrs"].sum()),2),
        )
    return vals

# ---------- cached-value injection (keeps formulas, adds <v> for any viewer) ----------
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
def inject_cache(path, sheet_filename, cache_by_ref):
    ET.register_namespace("", NS)
    tmp = path + ".zip"
    shutil.copy(path, tmp)
    with zipfile.ZipFile(tmp) as zin:
        names = zin.namelist()
        data = {n: zin.read(n) for n in names}
    xml = data[sheet_filename]
    root = ET.fromstring(xml)
    for c in root.iter(f"{{{NS}}}c"):
        ref = c.get("r")
        if ref not in cache_by_ref:        # only formula cells we computed
            continue
        f = c.find(f"{{{NS}}}f")
        if f is None:
            continue
        val = cache_by_ref[ref]
        for v in c.findall(f"{{{NS}}}v"):
            c.remove(v)
        v = ET.SubElement(c, f"{{{NS}}}v")
        if isinstance(val, str):
            c.set("t", "str");  v.text = val
        else:
            if "t" in c.attrib and c.get("t") == "str":
                del c.attrib["t"]
            v.text = repr(val) if isinstance(val,float) else str(val)
    data[sheet_filename] = ET.tostring(root, xml_declaration=True, encoding="UTF-8")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, data[n])
    os.remove(tmp)

def sheet_file_for(xlsx_path, sheet_name):
    with zipfile.ZipFile(xlsx_path) as z:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid = None
    for s in wb.iter(f"{{{NS}}}sheet"):
        if s.get("name") == sheet_name:
            rid = s.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
    RNS = "http://schemas.openxmlformats.org/package/2006/relationships"
    for rel in rels.iter(f"{{{RNS}}}Relationship"):
        if rel.get("Id") == rid:
            tgt = rel.get("Target")
            return "xl/" + tgt if not tgt.startswith("/") else tgt.lstrip("/")
    raise RuntimeError("sheet not found")

# ---------- main build ----------
def build(new_master_path, ref_path, out_path):
    new = read_any(new_master_path)
    new.columns = [str(c) for c in new.columns]
    new["date"] = pd.to_datetime(new["date"], errors="coerce")
    new["running_hrs"] = pd.to_numeric(new["running_hrs"], errors="coerce").fillna(0)
    new["status"] = new["status"].astype(str).str.strip()

    wb = load_workbook(ref_path)                 # template: keeps styles/CF/widths/header
    ms, ss = wb[MASTER_SHEET], wb[SUMMARY_SHEET]

    # styles to clone onto new rows
    master_styles  = {c: copy(ms.cell(2,c)._style) for c in range(1, ms.max_column+1)}
    summary_styles = {c: copy(ss.cell(2,c)._style) for c in range(1, 22)}
    master_headers = [ms.cell(1,c).value for c in range(1, ms.max_column+1)]

    # ----- rewrite Master -----
    if ms.max_row > 1: ms.delete_rows(2, ms.max_row-1)
    cols = [h for h in master_headers if h in new.columns]
    for c,h in enumerate(master_headers, start=1):
        ms.cell(1,c, h)
    for i,(_,row) in enumerate(new.iterrows(), start=2):
        for c,h in enumerate(master_headers, start=1):
            v = row[h] if h in new.columns else None
            if h == "date" and pd.notna(v): v = pd.Timestamp(v).to_pydatetime()
            ms.cell(i,c, None if (pd.isna(v) if not isinstance(v,(list,dict)) else False) else v)
            ms.cell(i,c)._style = copy(master_styles[c])
    N = len(new) + 1                              # Master last row index

    # ----- deployment list (one row per distinct asset+site) -----
    pairs = (new[["equipmentnumber","sitecode"]].dropna()
                .drop_duplicates().sort_values(["sitecode","equipmentnumber"]))
    pairs = list(pairs.itertuples(index=False, name=None))
    first_by_asset, site_meta = {}, {}
    for _,r in new.iterrows():
        a=r["equipmentnumber"]
        if a not in first_by_asset:
            first_by_asset[a]=(r.get("make",""), r.get("model",""), r.get("fleettype",""))
        key=(a, r["sitecode"])
        if key not in site_meta:
            site_meta[key]=(r.get("party_name",""), r.get("sitelocation",""))

    # ----- rewrite Summery report -----
    if ss.max_row > 1: ss.delete_rows(2, ss.max_row-1)
    cache = {}
    valmap = compute_values(new, pairs, first_by_asset)
    for idx,(asset,site) in enumerate(pairs, start=1):
        r = idx + 1
        fm = formulas(r, N)
        client, location = site_meta.get((asset,site), ("",""))
        # inputs
        ss.cell(r,2, asset); ss.cell(r,7, site)
        ss.cell(r,8, client); ss.cell(r,9, location)
        # formulas
        ss.cell(r,1,  fm["A"])
        ss.cell(r,3,  fm["C"]); ss.cell(r,4, fm["D"]); ss.cell(r,6, fm["F"])
        ss.cell(r,10, fm["J"]); ss.cell(r,11, fm["K"]); ss.cell(r,12, fm["L"])
        for col,key in [(13,"M"),(14,"N"),(15,"O"),(17,"Q"),(18,"R"),(19,"S")]:
            L=get_column_letter(col)
            ss.cell(r,col, ArrayFormula(ref=f"{L}{r}", text=fm[key]))
        ss.cell(r,16, fm["P"]); ss.cell(r,20, fm["T"]); ss.cell(r,21, fm["U"])
        # style clone + number formats preserved from row-2 template
        for c in range(1,22):
            ss.cell(r,c)._style = copy(summary_styles[c])
        # cached values
        v = valmap[idx]
        for col,key in [(1,"A"),(3,"C"),(4,"D"),(6,"F"),(10,"J"),(11,"K"),(12,"L"),
                        (13,"M"),(14,"N"),(15,"O"),(16,"P"),(17,"Q"),(18,"R"),
                        (19,"S"),(20,"T"),(21,"U")]:
            cv = v[key]
            if cv is not None:
                cache[f"{get_column_letter(col)}{r}"] = cv

    last = len(pairs) + 1
    ss.auto_filter.ref = f"A1:U{last}"
    wb.properties.title = "Vision Infra Summary Report"
    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)

    # inject cached values so the file is correct in every viewer (Excel still recalcs)
    inject_cache(out_path, sheet_file_for(out_path, SUMMARY_SHEET), cache)
    return len(new), len(pairs)

if __name__ == "__main__":
    new_master = sys.argv[1]
    ref  = sys.argv[2] if len(sys.argv) > 2 else "final_fine.xlsx"
    out  = sys.argv[3] if len(sys.argv) > 3 else "Summary-Report-Output.xlsx"
    nrows, prows = build(new_master, ref, out)
    print(f"Master rows: {nrows} | Summary deployment rows: {prows} -> {out}")