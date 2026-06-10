"""
Vision Infra - Rental Logsheet ETL
Transforms the raw daily logsheet dump (ERP export) into the monthly
per-asset PNM template format, applying the template's exact formulas.

Usage:  python etl_transform.py <dump.xls|.xlsx> <TEMPLATE-PNM.xlsx> <output.xlsx>
"""
import sys, re, subprocess, os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

def read_any(path):
    if path.lower().endswith('.xlsx'):
        return pd.read_excel(path)
    out_dir = os.path.dirname(os.path.abspath(path)) or '.'
    subprocess.run(['libreoffice','--headless','--convert-to','xlsx',
                    '--outdir', out_dir, path], check=True,
                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    xlsx = os.path.splitext(path)[0] + '.xlsx'
    return pd.read_excel(xlsx)

def yom(code):
    m = re.match(r'^[A-Za-z]{4}(\d{2})\d', str(code))
    return 2000 + int(m.group(1)) if m else None

def first_valid(series):
    s = series.dropna()
    return s.iloc[0] if len(s) else ''

def aggregate(df):
    df = df.copy()
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    df = df.dropna(subset=['equipmentnumber', 'date'])
    df['running_hrs'] = pd.to_numeric(df['running_hrs'], errors='coerce').fillna(0)
    df['status'] = df['status'].fillna('').astype(str).str.strip()
    df = df.drop_duplicates(
        subset=['doc_no', 'date', 'equipmentnumber', 'status', 'running_hrs', 'sitecode'])
    df['month'] = df['date'].dt.to_period('M')

    rows = []
    for (eq, site, party, month), sub in df.groupby(
            ['equipmentnumber', 'sitecode', 'party_name', 'month'], dropna=False):
        dates = sub.groupby('date')['status'].first()
        cnt = dates.value_counts()
        start, close = sub['date'].min(), sub['date'].max()
        rows.append({
            'Asset Code': eq,
            'Make': first_valid(sub['make']),
            'Model': first_valid(sub['model']),
            'YOM': yom(eq),
            'DEPARTMENT': str(first_valid(sub['fleettype'])).upper(),
            'SITE CODE': site,
            'CLIENT NAME': party,
            'LOCATION': first_valid(sub['sitelocation']),
            'MONTH/YEAR': str(month),
            'START DATE': start.date(),
            'CLOSE DATE': close.date(),
            'Total Days': (close - start).days + 1,
            'Free Days': int(cnt.get('Free Asset', 0)),
            'IN TRANSIT DAYS': int(cnt.get('Transit', 0)),
            'BDWN Days': int(cnt.get('Breakdown', 0)),
            'Idle Days': int(cnt.get('Idle', 0)),
            'Working Days': int(cnt.get('Working', 0)),
            'WORKED HOURS': round(sub.loc[sub['status'] == 'Working', 'running_hrs'].sum(), 2),
        })
    return pd.DataFrame(rows).sort_values(
        ['MONTH/YEAR', 'DEPARTMENT', 'Asset Code']).reset_index(drop=True)

COLMAP = {
    'Asset Code':'B','Make':'C','Model':'D','YOM':'E','DEPARTMENT':'F',
    'SITE CODE':'I','CLIENT NAME':'J','LOCATION':'K','MONTH/YEAR':'L',
    'START DATE':'M','CLOSE DATE':'N','Total Days':'O','Free Days':'P',
    'IN TRANSIT DAYS':'Q','BDWN Days':'S','Idle Days':'T','Working Days':'U',
    'WORKED HOURS':'W',
}

def write_output(agg, template_path, out_path):
    wb = load_workbook(template_path)
    ws = wb['Sheet1']
    style_row = 2
    styles = {c: copy(ws.cell(style_row, c)._style) for c in range(1, 33)}
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for i, rec in enumerate(agg.itertuples(index=False), start=2):
        d = dict(zip(agg.columns, rec))
        for col, letter in COLMAP.items():
            ws[f'{letter}{i}'] = d[col]
        ws[f'A{i}'] = f'=ROW(A{i-1})'
        ws[f'R{i}'] = f'=O{i}-P{i}-Q{i}'                 # Deployed = Total-Free-Transit
        ws[f'V{i}'] = f'=U{i}+T{i}'                      # Billing  = Working+Idle
        ws[f'X{i}'] = f'=260/30*V{i}'                    # WO-Hour  = 260/30 * Billing
        ws[f'Y{i}'] = f'=IFERROR(W{i}/X{i},0)'           # Util/hr  = Worked/WO-Hour
        ws[f'Z{i}'] = f'=IFERROR(V{i}/R{i},0)'           # Util/day = Billing/Deployed
        ws[f'AE{i}'] = f'=AD{i}+AC{i}+AB{i}'             # Total Cost
        ws[f'AF{i}'] = f'=IFERROR(AE{i}/W{i},0)'         # Cost/hour
        for c in range(1, 33):
            ws.cell(i, c)._style = copy(styles[c])
    wb.save(out_path)

def main():
    dump, template, out = sys.argv[1], sys.argv[2], sys.argv[3]
    df = read_any(dump)
    agg = aggregate(df)
    write_output(agg, template, out)
    print(f'Wrote {len(agg)} rows -> {out}')

if __name__ == '__main__':
    main()