#!/usr/bin/env python3
"""
Vision Infra - Combine split deployment rows (across months)
============================================================
Rows that share the SAME deployment identity:
    Asset Code, Make, Model, YOM, DEPARTMENT, SITE CODE, CLIENT NAME, LOCATION
are merged into ONE row - even when they fall in different months.

  Total Days / Free / Transit / BDWN / Idle / Working / Worked Hours = SUM
  START DATE = start of the row that contributes the most days (main period)
  CLOSE DATE = START + (Total Days - 1)        -> one continuous block
  MONTH/YEAR = month of START
  DEPLOYED   = Total - Free - Transit           (formula)
  Billing    = Idle + Working                   (formula)

Example (VTCM180009 @ PARNER):
  (30-04 -> 30-04, 1 day)  +  (01-05 -> 13-05, 13 days)
   = 01-05-2026 -> 14-05-2026, 14 days

Usage:
  python combine_deployments.py <input.xlsx|.csv|.tsv> [output.xlsx]
"""
import sys, re, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

norm = lambda s: re.sub(r'[^a-z0-9]', '', str(s).lower())

F = {
 'asset':['assetcode'],'make':['make'],'model':['model'],'yom':['yom'],
 'dept':['department','dept'],'site':['sitecode'],'client':['clientname'],
 'location':['location'],'month':['monthyear'],'start':['startdate'],'close':['closedate'],
 'total':['totaldays'],'free':['freedays'],'transit':['intransitdays','transitdays'],
 'bdwn':['bdwndays'],'idle':['idledays'],'working':['workingdays'],'worked':['workedhours'],
}
# identity that defines "the same deployment" (NOTE: no month, no dates)
KEY  = ['asset','make','model','yom','dept','site','client','location']
SUMC = ['total','free','transit','bdwn','idle','working','worked']

def load(path):
    if path.lower().endswith('.csv'): return pd.read_csv(path, dtype=str)
    if path.lower().endswith('.tsv'): return pd.read_csv(path, sep='\t', dtype=str)
    return pd.read_excel(path, dtype=str)

def colmap(cols):
    nmap = {norm(c): c for c in cols}
    out = {}
    for f, al in F.items():
        for a in al:
            if a in nmap: out[f] = nmap[a]; break
    return out

def to_date(v):
    v = str(v).strip()
    for fmt in ('%d-%m-%Y','%Y-%m-%d','%d/%m/%Y','%d-%b-%Y','%d-%b-%y','%m/%d/%Y'):
        try: return datetime.datetime.strptime(v, fmt).date()
        except ValueError: pass
    try: return pd.to_datetime(v, dayfirst=True).date()
    except Exception: return None

def num(v):
    try: return float(str(v).replace(',', '').strip())
    except Exception: return 0.0

def combine(df, m):
    g = lambda r, f: r[m[f]] if f in m else ''
    recs = []
    for _, r in df.iterrows():
        if not str(g(r, 'asset')).strip(): continue
        recs.append({**{k: str(g(r, k)).strip() for k in KEY},
                     'start': to_date(g(r, 'start')), 'close': to_date(g(r, 'close')),
                     **{c: num(g(r, c)) for c in SUMC}})
    work = pd.DataFrame(recs)

    out_rows = []
    for keyvals, sub in work.groupby(KEY, dropna=False, sort=False):
        d = dict(zip(KEY, keyvals if isinstance(keyvals, tuple) else (keyvals,)))
        for c in SUMC:
            d[c] = sub[c].sum()
        # START = start of the row with the most days (the main deployment period)
        main = sub.sort_values(['total'], ascending=False).iloc[0]
        start = main['start']
        if start is None:                                   # fallback to earliest known
            valid = sub['start'].dropna()
            start = min(valid) if len(valid) else None
        total = int(d['total'])
        close = (start + datetime.timedelta(days=total - 1)) if (start and total > 0) else start
        d['start'], d['close'] = start, close
        out_rows.append(d)
    out = pd.DataFrame(out_rows).sort_values(['site', 'asset']).reset_index(drop=True)
    return work, out

HEAD = ['Sr No','Asset Code','Make ','Model ','YOM','DEPARTMENT','SITE CODE','CLIENT NAME',
        'LOCATION','MONTH/YEAR','START DATE','CLOSE DATE','Total Days ','Free Days ',
        'IN TRANSIT DAYS','DEPLOYED DAYS','BDWN Days ','Idle  Days ','Working Days',
        'Billing Days ','WORKED HOURS']
WIDTHS = [6,12.7,12,18,6,12,38,42,16,12,14,14,11,10,16,16,12,10,14,12,14]

def write(out, path):
    wb = Workbook(); ws = wb.active; ws.title = 'Summery report'
    hf = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='305496'); ctr = Alignment('center','center',wrap_text=True)
    side = Side(style='thin', color='B9C4D4'); bd = Border(side, side, side, side)
    for c, h in enumerate(HEAD, 1):
        cell = ws.cell(1, c, h); cell.font = hf; cell.fill = fill; cell.alignment = ctr; cell.border = bd
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c-1]
    ws.freeze_panes = 'A2'
    for i, r in enumerate(out.itertuples(index=False), start=2):
        d = dict(zip(out.columns, r))
        ws.cell(i,1, '=ROW()-1')
        ws.cell(i,2, d['asset']); ws.cell(i,3, d['make']); ws.cell(i,4, d['model'])
        ws.cell(i,5, d['yom']); ws.cell(i,6, d['dept']); ws.cell(i,7, d['site'])
        ws.cell(i,8, d['client']); ws.cell(i,9, d['location'])
        ws.cell(i,10, d['start'].strftime('%b-%Y') if d['start'] else (d.get('month') or ''))
        cs = ws.cell(i,11, d['start']); ce = ws.cell(i,12, d['close'])
        cs.number_format = ce.number_format = 'dd-mmm-yyyy'
        ws.cell(i,13, int(d['total'])); ws.cell(i,14, int(d['free'])); ws.cell(i,15, int(d['transit']))
        ws.cell(i,16, f'=M{i}-N{i}-O{i}')
        ws.cell(i,17, int(d['bdwn'])); ws.cell(i,18, int(d['idle'])); ws.cell(i,19, int(d['working']))
        ws.cell(i,20, f'=R{i}+S{i}')
        wc = ws.cell(i,21, round(d['worked'],2)); wc.number_format = '0.00'
        for c in range(1,22): ws.cell(i,c).border = bd
    ws.auto_filter.ref = f'A1:U{out.shape[0]+1}'
    wb.calculation.fullCalcOnLoad = True
    wb.save(path)

if __name__ == '__main__':
    src = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else 'Combined-Summary-Report.xlsx'
    df = load(src); m = colmap(df.columns)
    work, comb = combine(df, m)
    write(comb, out_path)
    print(f'Input rows: {len(work)} | Combined rows: {len(comb)} | Collapsed: {len(work)-len(comb)} -> {out_path}')
