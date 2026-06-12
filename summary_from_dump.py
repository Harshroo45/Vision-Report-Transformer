#!/usr/bin/env python3
"""
Vision Infra - Summary report from raw logsheet (one row per Asset x Site x Month)
==================================================================================
Reads the raw daily logsheet dump and produces the 21-column summary. Each
distinct (Asset Code, Site Code, Month) becomes ONE row - so deployments that
were previously split across several rows are naturally combined.

  START DATE  = earliest log date in the group
  CLOSE DATE  = latest log date in the group
  Total Days  = distinct calendar dates logged
  Free/Transit/BDWN/Idle/Working Days = distinct dates with that status
  DEPLOYED    = Total - Free - Transit         (formula)
  WORKED HOURS= SUM of running_hrs for ALL statuses (matches the SUMIFS logic)

Usage:
  python summary_from_dump.py <dump.xlsx|.xls|.csv> [output.xlsx]
"""
import sys, os, subprocess, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def read_any(path):
    if path.lower().endswith('.csv'): return pd.read_csv(path)
    if path.lower().endswith('.xls'):
        out = os.path.dirname(os.path.abspath(path)) or '.'
        subprocess.run(['libreoffice','--headless','--convert-to','xlsx','--outdir',out,path],
                       check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        path = os.path.splitext(path)[0] + '.xlsx'
    xl = pd.ExcelFile(path)
    sheet = 'Master' if 'Master' in xl.sheet_names else xl.sheet_names[0]
    return pd.read_excel(path, sheet_name=sheet)

def first(s):
    s = s.dropna()
    return s.iloc[0] if len(s) else ''

def build(dump_path, out_path):
    m = read_any(dump_path)
    m.columns = [str(c) for c in m.columns]
    m['date'] = pd.to_datetime(m['date'], errors='coerce')
    m['running_hrs'] = pd.to_numeric(m['running_hrs'], errors='coerce').fillna(0)
    m['status'] = m['status'].astype(str).str.strip()
    m = m.dropna(subset=['equipmentnumber','date'])
    m['month'] = m['date'].dt.to_period('M')

    rows = []
    for (asset, site, month), sub in m.groupby(['equipmentnumber','sitecode','month'], dropna=False):
        dts = sub.groupby(sub['date'].dt.normalize())['status'].first()   # 1 status / date
        cnt = dts.value_counts()
        start, close = sub['date'].min(), sub['date'].max()
        rows.append(dict(
            asset=asset, make=first(sub['make']), model=first(sub['model']), yom='',
            dept=str(first(sub['fleettype'])).upper(), site=site,
            client=first(sub['party_name']), location=first(sub['sitelocation']),
            month=month.strftime('%b-%Y'), start=start.date(), close=close.date(),
            total=int(dts.shape[0]),
            free=int(cnt.get('Free Asset',0)), transit=int(cnt.get('Transit',0)),
            bdwn=int(cnt.get('Breakdown',0)), idle=int(cnt.get('Idle',0)),
            working=int(cnt.get('Working',0)),
            worked=round(float(sub['running_hrs'].sum()),2),   # ALL running hours
        ))
    out = pd.DataFrame(rows).sort_values(['month','site','asset']).reset_index(drop=True)
    write(out, out_path)
    return len(m), len(out)

HEAD = ['Sr No','Asset Code','Make ','Model ','YOM','DEPARTMENT','SITE CODE','CLIENT NAME',
        'LOCATION','MONTH/YEAR','START DATE','CLOSE DATE','Total Days ','Free Days ',
        'IN TRANSIT DAYS','DEPLOYED DAYS','BDWN Days ','Idle  Days ','Working Days',
        'WORKED HOURS']
WIDTHS = [6,12.7,12,18,6,12,38,42,16,12,14,14,11,10,16,16,12,10,14,14]

def write(out, path):
    wb = Workbook(); ws = wb.active; ws.title = 'Summery report'
    hf = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='305496'); ctr = Alignment('center','center',wrap_text=True)
    side = Side(style='thin', color='B9C4D4'); bd = Border(side,side,side,side)
    for c,h in enumerate(HEAD,1):
        cell = ws.cell(1,c,h); cell.font=hf; cell.fill=fill; cell.alignment=ctr; cell.border=bd
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c-1]
    ws.freeze_panes = 'A2'
    for i, r in enumerate(out.itertuples(index=False), start=2):
        d = dict(zip(out.columns, r))
        ws.cell(i,1, '=ROW()-1')
        ws.cell(i,2, d['asset']); ws.cell(i,3, d['make']); ws.cell(i,4, d['model'])
        ws.cell(i,5, d['yom']); ws.cell(i,6, d['dept']); ws.cell(i,7, d['site'])
        ws.cell(i,8, d['client']); ws.cell(i,9, d['location']); ws.cell(i,10, d['month'])
        cs=ws.cell(i,11, d['start']); ce=ws.cell(i,12, d['close']); cs.number_format=ce.number_format='dd-mmm-yyyy'
        ws.cell(i,13, d['total']); ws.cell(i,14, d['free']); ws.cell(i,15, d['transit'])
        ws.cell(i,16, f'=M{i}-N{i}-O{i}')
        ws.cell(i,17, d['bdwn']); ws.cell(i,18, d['idle']); ws.cell(i,19, d['working'])
        wc=ws.cell(i,20, d['worked']); wc.number_format='0.00'
        for c in range(1,21): ws.cell(i,c).border = bd
    ws.auto_filter.ref = f'A1:T{out.shape[0]+1}'
    wb.calculation.fullCalcOnLoad = True
    wb.save(path)

if __name__ == '__main__':
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else 'Summary-Report-Merged.xlsx'
    nrows, prows = build(src, out)
    print(f'Logsheet rows: {nrows} | Summary rows (Asset x Site x Month): {prows} -> {out}')
